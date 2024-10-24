import pandas as pd
import numpy as np
from openpyxl import load_workbook

def extract_hyperlinks(wb, sheet_name):
    """
    Extract hyperlinks from a worksheet using a method compatible with read-only mode
    """
    ws = wb[sheet_name]
    links = []
    
    # Get the column index for 'Proposal' (usually A/1)
    proposal_col = 1  # Assuming 'Proposal' is always in column A
    
    # Skip header row
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=proposal_col)
        if cell.hyperlink:
            links.append(cell.hyperlink.target)
        else:
            links.append(np.nan)
    
    return links

def process_excel_file(input_file, output_file, fund_name, skip_sheets):
    """
    Process Excel file with voting results and export to CSV.
    
    Parameters:
    input_file (str): Path to input Excel file
    output_file (str): Path to output CSV file
    fund_name (str): Name of the fund (e.g., 'Fund11')
    skip_sheets (list): List of worksheet names to skip
    """
    
    # Read all sheets except those in skip_sheets
    excel_file = pd.ExcelFile(input_file)
    valid_sheets = [sheet for sheet in excel_file.sheet_names if sheet not in skip_sheets]
    
    # Load workbook for hyperlink extraction
    wb = load_workbook(input_file, read_only=False)  # Temporarily disable read_only mode
    
    # Initialize list to store DataFrames
    dfs = []
    
    for sheet in valid_sheets:
        # Read the sheet
        df = pd.read_excel(
            input_file,
            sheet_name=sheet,
            engine='openpyxl'
        )
        
        # Extract hyperlinks
        try:
            links = extract_hyperlinks(wb, sheet)
            
            # Ensure the number of links matches the DataFrame length
            if len(links) == len(df):
                df['Link'] = links
            else:
                print(f"Warning: Number of extracted links ({len(links)}) doesn't match data rows ({len(df)}) in sheet {sheet}")
                df['Link'] = np.nan
                
        except Exception as e:
            print(f"Warning: Could not extract hyperlinks from sheet {sheet}: {str(e)}")
            df['Link'] = np.nan
        
        # Clean up the Proposal column (remove line breaks but keep special characters)
        df['Proposal'] = df['Proposal'].astype(str).apply(lambda x: ' '.join(x.splitlines()))
        
        # Add fund name and sheet name columns
        df['Fund'] = fund_name
        df['Category'] = sheet
        
        dfs.append(df)
    
    # Close the workbook
    wb.close()
    
    # Combine all DataFrames
    final_df = pd.concat(dfs, ignore_index=True)
    
    # Export to CSV
    final_df.to_csv(output_file, index=False, encoding='utf-8')
    
    print(f"Processing complete. Data exported to {output_file}")
    return final_df

# Example usage
if __name__ == "__main__":
    input_file = "Fund11 Voting Results.xlsx"
    output_file = "Fund11_voting_data.csv"
    fund_name = "Fund11"
    skip_sheets = ["Sponsored by leftovers", "Validation"]
    
    df = process_excel_file(input_file, output_file, fund_name, skip_sheets)
